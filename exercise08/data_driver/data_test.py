# coding=utf-8
'''
@Time   :2020/12/4 15:44
@Author :六月
@Email  :juneren26@gmail.com
@File   :data_test.py
@IDE    :PyCharm
'''
from openpyxl.styles import PatternFill, Font

'''
    基于excel数据驱动来实现自动化测试
    封装好关键字后，通过读取数据，让他自行执行对应的关键字对象
'''
import openpyxl
# 访问excel内容
from exercise07.key_word_web.keyword_case import WebKeys
from logs.log import Logger
# 日志
log = Logger().get_logger()
excel = openpyxl.load_workbook('../data/xxx.xlsx')
sheet = excel['Sheet3']
log.info('获取{}内容成功，现在开始执行自动化测试......'.format(sheet))
# 读取excel内容，实现文件驱动自动化执行
for value in sheet.values:
    # 定义一个字典参数接收excel中的全部数据
    args = {}
    # 定位方法
    args['name'] = value[2]
    # 定位路径
    args['value'] = value[3]
    # 输入内容
    args['txt'] = value[4]
    # 预期结果
    args['expect'] = value[6]
    # 基于A列进行判断是否为测试用例
    if type(value[0]) is int:
        '''
        在读取关键字时，分为几类情况：
            1.关键字驱动类的实例化
            2.断言类型的关键字W
        '''
        # 判断是否实例化
        if value[1] == 'open_browser':
            log.info('现在执行关键字：{},操作描述：{}'.format(value[1],value[5]))
            wk = WebKeys(value[4])
        # 断言关键字执行：首先执行断言，在判断断言是否成功
        elif 'assert' in value[1]:
            log.info('现在执行关键字:{0}，操作描述：{1}'.format(value[1], value[5]))
            status = getattr(wk,value[1])(**args)
            if status:
                # 写入当前行的实际结果值
                # sheet.cell(row='编号+1',column=8)
                sheet.cell(row=value[0] + 1, column=8).value = 'Pass'
                sheet.cell(row=value[0] + 1, column=8).fill = PatternFill('solid', fgColor='AACF91')
                sheet.cell(row=value[0] + 1, column=8).font = Font(bold=True)
            else:
                sheet.cell(row=value[0] + 1, column=8).value = 'Failed'
                sheet.cell(row=value[0] + 1, column=8).fill = PatternFill('solid', fgColor='FF0000')
                sheet.cell(row=value[0] + 1, column=8).font = Font(bold=True)
            excel.save('../data/xxx.xlsx')
            # 依据status来判定写入的内容是Failed还是Pass
        # 非特殊关键字，通过反射机制实现,反射机制接收两个参数（第一个参数为对象或者模块，第二个参数是一个字符串）
        else:
            log.info('现在执行关键字:{},操作的描述：{}'.format(value[1],value[5]))
            getattr(wk, value[1])(**args)
            # getattr(wk,'open')(**args)
            # wk.open(**args)
            # 原有的open函数是 def open(self,url)
excel.close()